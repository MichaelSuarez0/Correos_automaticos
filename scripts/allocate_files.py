## 0. Variables
from correos_automaticos.classes.sharepoint_manager import Sharepoint
from correos_automaticos.classes.file_manager import FileManager
from dotenv import load_dotenv
import os
import pandas as pd
import urllib.parse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

script_dir = os.path.dirname(__file__) # for .py files
#script_dir = os.getcwd()  # for jupyter

# Variables globales y de entorno
load_dotenv()
DOWNLOAD_PATH = os.path.join(script_dir, "..", "descargas")  # Carpeta de descargas
UPLOAD_PATH = os.path.join(script_dir, "..", "descargas", "clasificados")  # Carpeta desde donde se subirán archivos
TEMPLATES_PATH = os.path.join(script_dir, "..", "email_templates") # Carpeta desde la que se obtendrán los email templates

## 1. Iniciar variables
file_manager = FileManager(search_directory= UPLOAD_PATH)
excel_path = os.path.join(script_dir, "..", "docs", "Registro de Participación con adjuntos_v5.xlsx")
df_merged = pd.read_excel(excel_path)
df_merged = df_merged.drop(columns=["Id", "Adjuntar programa", "full_url", "time_created", "Hora de inicio", "Hora de finalización"], axis=1)

meses = {
    "01": "01 - Enero", "02": "02 - Febrero", "03": "03 - Marzo", "04": "04 - Abril",
    "05": "05 - Mayo", "06": "06 - Junio", "07": "07 - Julio", "08": "08 - Agosto",
    "09": "09 - Septiembre", "10": "10 - Octubre", "11": "11 - Noviembre", "12": "12 - Diciembre"
}

peru_sectors = {
    "Agricultura y Riego": "MINAGRI",
    "Ambiente": "MINAM",
    "Cultura": "MINCUL",
    "Defensa": "MINDEF",
    "Desarrollo e Inclusión Social": "MIDIS",
    "Comercio Exterior y Turismo": "MINCETUR",
    "Economía y Finanzas": "MEF",
    "Educación": "MINEDU",
    "Energía y Minas": "MINEM",
    "Interior": "MININTER",
    "Gobernación": "MININTER",
    "Justicia y Derechos Humanos": "MINJUSDH",
    "Mujer y Poblaciones Vulnerables": "MIMP",
    "Salud": "MINSA",
    "Vivienda, Construcción y Saneamiento": "MVCS",
    "Presidencia Consejo de Ministros (PCM)": "PCM",
    "Producción": "PRODUCE",
    "Relaciones Exteriores": "RREE",
    "Trabajo": "MTPE",
    "Transportes y Comunicaciones": "MTC"
}

peru_departamentos = {
    "Amazonas": "AMAZONAS",
    "Áncash": "ANCASH",
    "Apurímac": "APURIMAC",
    "Arequipa": "AREQUIPA",
    "Ayacucho": "AYACUCHO",
    "Cajamarca": "CAJAMARCA",
    "Callao": "CALLAO",
    "Cusco": "CUSCO",
    "Huancavelica": "HUANCAVELICA",
    "Huánuco": "HUANUCO",
    "Ica": "ICA",
    "Junín": "JUNIN",
    "La Libertad": "LIBERTAD",
    "Lambayeque": "LAMBAYEQUE",
    "Lima Metropolitana": "LIMAMETR",
    "Lima": "LIMAREGION",
    "Lima Provincias": "LIMAREGION",
    "Loreto": "LORETO",
    "Madre de Dios": "MADREDEDIOS",
    "Moquegua": "MOQUEGUA",
    "Pasco": "PASCO",
    "Piura": "PIURA",
    "Puno": "PUNO",
    "San Martín": "SANMARTIN",
    "Tacna": "TACNA",
    "Tumbes": "TUMBES",
    "Ucayali": "UCAYALI"
}

columnas_politicas = [
    "Actividad operativa", "Especialista de la DNPE a cargo", "Fecha de ejecución de la actividad",
    "Nombre de la política / plan", "Naturaleza del trabajo", "Nivel de Gobierno", "Poder del Estado", "Sector",
    "Entidad (OCA o institución estatal)", "Región", "Etapa/fase de acompañamiento",
    "Estado del proceso", "Dirección de la persona de contacto/coordinación"
]

columnas_talleres = [
    "Actividad operativa", "Especialista de la DNPE a cargo", "Fecha de ejecución de la actividad",
    "Nombre del taller / evento", "Naturaleza del trabajo","Nivel de Gobierno", "Poder del Estado", "Sector",
    "Entidad (OCA o institución estatal)", "Región", "Dirección de la persona de contacto/coordinación"
]

oca = [
    "OCA",
    "Organismo autónomo",
    "Organismo Constitucional Autónomo",
    "Organismo Constitucionalmente Autónoma",
    "Organismo Constitucionalmente Autónomo",
    "Organismos Constitucionales Autónomos"
]


## 2. Definir funciones principales
def construct_code(data, file_name: str):
    code = ""
    if file_name in data["name"].values:
        row_index = data.index[data["name"] == file_name][0]

        # Actividad operativa
        AOI = data.loc[row_index, "Actividad operativa"]
        if AOI == "Espacios de difusión (Estudios/plataformas)":
            AOI = "Espacios de difusión (Estudios y plataformas)" # Para que no haya problema con los paths
            code = "DIFUSION"
        elif AOI == "Convenios":
            AOI = "Espacios de consulta"
            code = "CONSULTA"
        elif AOI == "Asistencia técnica (Políticas y planes)":
            code = "ATECNICA"
        elif AOI == "Instrumentos técnicos en prospectiva":
            code = "INSTRUME"
        
        
        # Fecha 
        fecha = data.loc[row_index, "Fecha de ejecución de la actividad"]
        y,m,d = str(fecha.year), str(fecha.month).zfill(2), str(fecha.day).zfill(2)
        nombre_mes = meses[m]
        code = f'{code}-{d}-{m}'


        # Nivel de Gobierno
        nivel_gob = data.loc[row_index, "Nivel de Gobierno"]
        naturaleza = data.loc[row_index, "Naturaleza del trabajo"]
        if naturaleza == "Revisión de entregables":
            if nivel_gob == "Gobierno Nacional":
                code = f'{code}-PNAC'
            elif nivel_gob == "Gobierno Regional":
                code = f'{code}-PDRC'
            elif nivel_gob == "Gobierno Local":
                code = f'{code}-PDLC'
            else:
                code = f'{code}-OTRO'
        else:
            if nivel_gob == "Gobierno Nacional":
                code = f'{code}-GN'
            elif nivel_gob == "Gobierno Regional":
                code = f'{code}-GR'
            elif nivel_gob == "Gobierno Local":
                code = f'{code}-GL'
            elif nivel_gob in oca:
                code = f'{code}-OCA'
            else:
                code = f'{code}-OTRO'

        # Region o sector
        sector = peru_sectors.get(data.loc[row_index, "Sector"])
        region = peru_departamentos.get(data.loc[row_index, "Región"])
        if nivel_gob == "Gobierno Nacional":
            if sector is not None:
                code = f'{code}-{sector}'
            elif region is not None:
                code = f'{code}-{region}'
            else:
                code = f'{code}-OTRO'
        elif nivel_gob in ["Gobierno Regional", "Gobierno Local"] or nivel_gob in oca:
            if region is not None:
                code = f'{code}-{region}'
            elif sector is not None:
                code = f'{code}-{sector}'
            else:
                code = f'{code}-OTRO'
        else:
            code = f'{code}-OTRO'
        
        
        # # Naturaleza del trabajo
        # naturaleza = data.loc[row_index, "Naturaleza del trabajo"]
        # if naturaleza == "Revisión de entregables":
        #     code = f'{code}-ENTREG'
        # elif naturaleza in  ["Talleres", "Talleres de capacitación"]:
        #     code = f'{code}-TALLER'
        # elif naturaleza == "Webinar":
        #     code = f'{code}-WEBINR'
        # elif naturaleza == "Convenios":
        #     code = f'{code}-CONVEN'


        # Iniciales del autor
        autor = data.loc[row_index, "Especialista de la DNPE a cargo"]
        if autor in ["Enrique Del Águila", "Alberto Del Aguila"]:
            code = f'{code}-ADA'
        if autor == "Álvaro Gamboa":
            code = f'{code}-AGB'
        if autor == "Carmen Bahamonde":
            code = f'{code}-CBQ'
        if autor == "Eduardo Sobrino":
            code = f'{code}-ESV'
        if autor == "Erika Céliz":
            code = f'{code}-ECY'
        if autor == "John Pichihua":
            code = f'{code}-JPT'
        if autor == "Katherine Guadalupe":
            code = f'{code}-KGM'
        if autor == "Marco Francisco":
            code = f'{code}-MFT'
        if autor == "Milagros Estrada":
            code = f'{code}-MER'
        if autor == "Yiem Ataucusi":
            code = f'{code}-YAA'
        # else:
        #     autor = autor.split()
        #     inicial, segundo = autor[0], autor[1]
        #     code = f'{code}-{inicial[:1]}{segundo[:1]}'
        
        constructed_url = f'Documentos compartidos/AOI Asistencia técnica/Prueba/{AOI}/{nombre_mes}/{code}'
        #print(f' - URL: {constructed_url}')
        #print(f' - {code}')

        return constructed_url, code, row_index


def generar_metadata(data: pd.DataFrame, code: str, row_index: str) -> str :
    # Obtener datos preliminares
    # if file_name in data["name"].values:
    #     row_index = data.index[data["name"] == file_name][0]
    # Inicializar variables
    excel_file = f'{code}.xlsx'
    output_path = os.path.join(UPLOAD_PATH, excel_file)
    naturaleza = data.loc[row_index, "Naturaleza del trabajo"]
    taller = False if naturaleza == "Revisión de entregables" else True
    if taller == True:
        important_columns = columnas_talleres
    else:
        important_columns = columnas_politicas
    data = data.drop(columns=["name"], axis=1)

    # Generar tablas de datos importantes y detalles
    important_data = {col: data.loc[row_index, col] for col in important_columns}
    detail_data = {col: data.loc[row_index, col] for col in data.columns if col not in important_columns}
    
    # Agregar el código asignado
    important_data["Código Asignado"] = code

    # Convertir tablas a DataFrames y transponerlos
    important_df = pd.DataFrame.from_dict(important_data, orient='index', columns=["Detalle"]).reset_index()
    important_df.columns = ["Pregunta", "Detalle"] 
    detail_df = pd.DataFrame.from_dict(detail_data, orient='index', columns=["Detalle"]).reset_index()
    detail_df.columns = ["Pregunta", "Detalle"]
    detail_df = detail_df.dropna()

    # Concatenar ambos DataFrames
    metadata_df = pd.concat([important_df, detail_df], ignore_index=True)

    # Guardar el DataFrame directamente al archivo Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        metadata_df.to_excel(writer, index=False, sheet_name="metadata")
        
        # Ajustar el ancho de la columnas
        worksheet = writer.sheets["metadata"]
        for column in metadata_df.columns:
            col_idx = metadata_df.columns.get_loc(column) + 1  # Obtener el índice de la columna (1-based)
            column_letter = chr(64 + col_idx)  # Convertir el índice a letra de columna
            worksheet.column_dimensions[column_letter].width = 40  # Establecer el ancho de la columna

            # Aplicar el ajuste de texto a cada celda de la columna
            for cell in worksheet[column_letter]:
                cell.alignment = Alignment(wrap_text=True)  # Ajustar texto. TODO: no funciona en excel en línea

        # Insertar la fila y configurar el encabezado "Información general"
        worksheet.merge_cells('A1:B1')
        worksheet['A1'] = 'Información general del evento'
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Determinar la fila en función de "taller"
        row = 15 if not taller else 13

        # Insertar fila y configurar el encabezado "Detalles"
        worksheet.insert_rows(row)
        worksheet.merge_cells(f'A{row}:B{row}')
        worksheet[f'A{row}'] = 'Detalles adicionales del evento'
        worksheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Aplicar color de fondo y fuente a los encabezados
        dark_blue_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        white_font = Font(color="FFFFFF")
        worksheet['A1'].fill = dark_blue_fill
        worksheet['A1'].font = white_font
        worksheet[f'A{row}'].fill = dark_blue_fill
        worksheet[f'A{row}'].font = white_font

        # Alinear el texto de la columna "Detalle" a la izquierda
        detalle_column_letter = 'B'  # La columna "Detalle" es la segunda columna (B)
        for cell in worksheet[detalle_column_letter]:
            cell.alignment = Alignment(horizontal='left', wrap_text=True)  # Alinear a la izquierda

        print(f" - Archivo de metadatos generado para {code}")
        return excel_file
    

def allocate_files_from_folder():
    session = Sharepoint()
    session._auth()
    file_list = file_manager.list_files()
    #session.download_files_from_folder()
    for file_name in file_list:
        try:
            constructed_url, code, row_index = construct_code(df_merged, file_name)
            excel_file = generar_metadata(df_merged, code, row_index)    
            session.upload_file(file_name=file_name, custom_folder_path=constructed_url, create_folder=True)
            session.upload_file(file_name=excel_file, custom_folder_path=constructed_url, create_folder=True)
        except Exception as e:
            print(f'ERROR: Hubo un problema con la subida del archivo "{file_name}": {e}')
 

if __name__ == "__main__":
    allocate_files_from_folder()
